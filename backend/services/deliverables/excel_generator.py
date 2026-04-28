"""Excel BOM generator — bảng vật tư có công thức tự động.

Output XLSX với:
- Sheet 1 "BOQ" — bảng vật tư + công thức =qty*unit_price
- Sheet 2 "Summary" — tổng hợp theo category với SUMIF formula
- Sheet 3 "Project" — thông tin dự án (prompt, style, area, variants)

User edit qty/unit_price → total_price tự cập nhật.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# ─── Color theme ──────────────────────────────────────────────
TEAL = "FF00C9A7"
DARK = "FF0A1020"
LIGHT = "FFF8F9FF"
GREY = "FF5A7898"
BORDER = "FFCDD8E5"


def _safe(v: Any, d: str = "") -> str:
    if v is None:
        return d
    s = str(v).strip()
    return s if s else d


def _border_thin() -> Border:
    side = Side(border_style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _hdr_cell(value: str) -> dict:
    return {
        "value": value,
        "font": Font(name="Calibri", size=11, bold=True, color="FFFFFFFF"),
        "fill": PatternFill(fill_type="solid", fgColor=TEAL),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _border_thin(),
    }


def _apply(cell, props: dict) -> None:
    if "value" in props:
        cell.value = props["value"]
    if "font" in props:
        cell.font = props["font"]
    if "fill" in props:
        cell.fill = props["fill"]
    if "alignment" in props:
        cell.alignment = props["alignment"]
    if "border" in props:
        cell.border = props["border"]
    if "number_format" in props:
        cell.number_format = props["number_format"]


# ─── Main entrypoint ──────────────────────────────────────────
def generate_bom_xlsx(design: dict, user: dict | None = None) -> bytes:
    """Generate Excel BOM file for a design.

    Returns: XLSX bytes ready for HTTP response.
    """
    wb = Workbook()

    # ─── Sheet 1: BOQ ─────────────────────────────────────────
    ws = wb.active
    ws.title = "BOQ"
    ws.sheet_view.zoomScale = 110

    # Column widths
    widths = [4, 18, 32, 10, 8, 14, 14, 12]
    headers = ["STT", "Hạng mục", "Vật liệu / Sản phẩm", "Số lượng", "ĐVT",
               "Đơn giá (VND)", "Thành tiền (VND)", "Trạng thái"]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = "BẢNG KHỐI LƯỢNG VẬT TƯ — NEXBUILD"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Project info row
    ws.merge_cells("A2:H2")
    info = (
        f"Phong cách: {_safe(design.get('style'),'—')} · "
        f"Diện tích: {_safe(design.get('area_m2'),'—')}m² · "
        f"Hạng mục: {_safe(design.get('room_type'),'—')} · "
        f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y')}"
    )
    ws["A2"] = info
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=GREY)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Header row
    HEADER_ROW = 4
    for col, h in enumerate(headers, 1):
        _apply(ws.cell(row=HEADER_ROW, column=col), _hdr_cell(h))
    ws.row_dimensions[HEADER_ROW].height = 30

    # Data rows
    boq = (
        design.get("boq_items")
        or design.get("agent_output", {}).get("boq_structural")
        or []
    )

    body_font = Font(name="Calibri", size=10)
    border = _border_thin()
    alt_fill = PatternFill(fill_type="solid", fgColor=LIGHT)

    for i, item in enumerate(boq, 1):
        r = HEADER_ROW + i
        qty = item.get("quantity") or item.get("qty") or 0
        try:
            qty = float(qty)
        except (ValueError, TypeError):
            qty = 0
        unit_price = int(item.get("unit_price") or 0)

        rows_data = [
            (1, i, "center"),
            (2, _safe(item.get("category") or item.get("section"), "—"), "left"),
            (3, _safe(item.get("product_name") or item.get("item"), "—"), "left"),
            (4, qty, "right"),
            (5, _safe(item.get("unit"), ""), "center"),
            (6, unit_price, "right"),
            # Formula: =D{r}*F{r}
            (7, f"=D{r}*F{r}", "right"),
            (8, _safe(item.get("order_status"), "Chưa đặt"), "center"),
        ]
        for col, val, align in rows_data:
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = body_font
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border
            if col in (6, 7):
                cell.number_format = '#,##0;[Red]-#,##0'
            if col == 4:
                cell.number_format = '#,##0.00'
            # Alternate row color
            if i % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[r].height = 22

    # Totals
    if boq:
        n = len(boq)
        first_data_row = HEADER_ROW + 1
        last_data_row = HEADER_ROW + n
        total_row = last_data_row + 2

        ws.merge_cells(f"A{total_row}:F{total_row}")
        c = ws.cell(row=total_row, column=1, value="CỘNG")
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        c.fill = PatternFill(fill_type="solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = border

        cell_subtotal = ws.cell(
            row=total_row, column=7,
            value=f"=SUM(G{first_data_row}:G{last_data_row})",
        )
        cell_subtotal.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        cell_subtotal.fill = PatternFill(fill_type="solid", fgColor=DARK)
        cell_subtotal.alignment = Alignment(horizontal="right", vertical="center")
        cell_subtotal.number_format = '#,##0'
        cell_subtotal.border = border

        # VAT row
        vat_row = total_row + 1
        ws.merge_cells(f"A{vat_row}:F{vat_row}")
        c = ws.cell(row=vat_row, column=1, value="VAT 10%")
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = border

        cell_vat = ws.cell(row=vat_row, column=7, value=f"=G{total_row}*0.1")
        cell_vat.font = Font(name="Calibri", size=10)
        cell_vat.alignment = Alignment(horizontal="right", vertical="center")
        cell_vat.number_format = '#,##0'
        cell_vat.border = border

        # Grand total row
        grand_row = vat_row + 1
        ws.merge_cells(f"A{grand_row}:F{grand_row}")
        c = ws.cell(row=grand_row, column=1, value="TỔNG (ĐÃ VAT)")
        c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
        c.fill = PatternFill(fill_type="solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = border

        cell_grand = ws.cell(row=grand_row, column=7, value=f"=G{total_row}+G{vat_row}")
        cell_grand.font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
        cell_grand.fill = PatternFill(fill_type="solid", fgColor=TEAL)
        cell_grand.alignment = Alignment(horizontal="right", vertical="center")
        cell_grand.number_format = '#,##0'
        cell_grand.border = border

    # Freeze top rows
    ws.freeze_panes = ws[f"A{HEADER_ROW + 1}"]

    # ─── Sheet 2: Summary by category (SUMIF) ─────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.zoomScale = 110
    for col, w in enumerate([4, 26, 16, 14], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "TỔNG HỢP THEO HẠNG MỤC"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color=DARK)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 26

    sum_headers = ["STT", "Hạng mục", "Số dòng", "Thành tiền (VND)"]
    for col, h in enumerate(sum_headers, 1):
        _apply(ws2.cell(row=3, column=col), _hdr_cell(h))
    ws2.row_dimensions[3].height = 28

    if boq:
        # Get unique categories
        seen_cats: list[str] = []
        for item in boq:
            cat = _safe(item.get("category") or item.get("section"), "Khác")
            if cat not in seen_cats:
                seen_cats.append(cat)

        for i, cat in enumerate(seen_cats, 1):
            r = 3 + i
            ws2.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
            ws2.cell(row=r, column=2, value=cat)
            ws2.cell(row=r, column=3, value=f'=COUNTIF(BOQ!B:B,"{cat}")')
            cell = ws2.cell(
                row=r, column=4,
                value=f'=SUMIF(BOQ!B:B,"{cat}",BOQ!G:G)',
            )
            cell.number_format = '#,##0'
            for col in range(1, 5):
                ws2.cell(row=r, column=col).border = border
                ws2.cell(row=r, column=col).font = body_font

    # ─── Sheet 3: Project info ──────────────────────────────────
    ws3 = wb.create_sheet("Project")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 60

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "THÔNG TIN DỰ ÁN"
    ws3["A1"].font = Font(name="Calibri", size=14, bold=True, color=DARK)
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 26

    project_data = [
        ("Mã thiết kế:", _safe(design.get("design_id"))),
        ("Khách hàng:", _safe((user or {}).get("full_name"), "—")),
        ("Email:", _safe((user or {}).get("email"), "—")),
        ("Hạng mục:", _safe(design.get("room_type"), "Thiết kế nội thất")),
        ("Phong cách:", _safe(design.get("style"))),
        ("Diện tích:", f"{_safe(design.get('area_m2'),'—')} m²"),
        ("Ngân sách:", f"{_safe(design.get('budget_million'),'—')} triệu VND"),
        ("Mô tả:", _safe(design.get("prompt"))[:500]),
        ("Ngày tạo:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Phương án:", str(len(design.get("variants", [])))),
    ]
    label_font = Font(name="Calibri", size=10, bold=True, color=GREY)
    value_font = Font(name="Calibri", size=10, color=DARK)

    for i, (label, value) in enumerate(project_data, 3):
        a = ws3.cell(row=i, column=1, value=label)
        a.font = label_font
        a.alignment = Alignment(horizontal="right", vertical="top")
        b = ws3.cell(row=i, column=2, value=value)
        b.font = value_font
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws3.row_dimensions[i].height = 20

    # Variants subsection
    variants = design.get("variants") or []
    if variants:
        var_start = len(project_data) + 5
        ws3.cell(row=var_start, column=1, value="Phương án:").font = label_font
        ws3.cell(row=var_start, column=1).alignment = Alignment(horizontal="right")
        for i, v in enumerate(variants[:4], 1):
            r = var_start + i
            label = _safe(v.get("style_label") or v.get("concept_name"), f"Phương án {i}")
            desc = _safe(v.get("description"))[:200]
            ws3.cell(row=r, column=1, value=f"• {label}").font = label_font
            ws3.cell(row=r, column=2, value=desc).font = value_font
            ws3.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws3.row_dimensions[r].height = 30

    # ─── Output bytes ───────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
